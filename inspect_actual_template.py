from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def value_text(value):
    if value is None:
        return ""
    return str(value).strip()


def row_values(row):
    return [value_text(cell) for cell in row]


def non_empty_count(values):
    return sum(1 for value in values if value)


def inspect(path: str):
    workbook = load_workbook(path, data_only=True, read_only=True)
    result = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = []
        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = row_values(row)
            if non_empty_count(values):
                rows.append({"row": index, "values": values[:40]})
            if len(rows) >= 12:
                break
        header_candidates = []
        for item in rows[:8]:
            values = item["values"]
            joined = " ".join(values)
            score = sum(keyword in joined for keyword in ["营期", "渠道", "分类", "GMV", "gmv", "Leads", "leads", "转化", "ROI", "消耗", "正价"])
            header_candidates.append({"row": item["row"], "score": score, "values": values[:40]})
        result[sheet_name] = {
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "samples": rows,
            "header_candidates": sorted(header_candidates, key=lambda x: x["score"], reverse=True)[:3],
        }
    return result


if __name__ == "__main__":
    payload = inspect(sys.argv[1])
    print(json.dumps(payload, ensure_ascii=False, indent=2))
