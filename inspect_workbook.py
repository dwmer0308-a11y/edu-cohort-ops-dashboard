from openpyxl import load_workbook
import json

path = "/Users/zhangliang/Downloads/【排量】书法项目-排期&营收-规划表.xlsx"
wb_values = load_workbook(path, data_only=True, read_only=False)
wb_formulas = load_workbook(path, data_only=False, read_only=False)

summary = []
for ws in wb_values.worksheets:
    non_empty = 0
    formula_count = 0
    max_row = ws.max_row
    max_col = ws.max_column
    fws = wb_formulas[ws.title]
    samples = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                non_empty += 1
    for row in fws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_count += 1
    for r in range(1, min(max_row, 30) + 1):
        values = []
        for c in range(1, min(max_col, 18) + 1):
            v = ws.cell(r, c).value
            if v is None:
                values.append("")
            elif hasattr(v, "isoformat"):
                values.append(v.isoformat())
            else:
                values.append(v)
        if any(x != "" for x in values):
            samples.append(values)
    summary.append({
        "sheet": ws.title,
        "max_row": max_row,
        "max_col": max_col,
        "non_empty_cells": non_empty,
        "formula_count": formula_count,
        "merged_ranges": [str(x) for x in list(ws.merged_cells.ranges)[:20]],
        "sample_rows": samples[:18],
    })

print(json.dumps(summary, ensure_ascii=False, indent=2))
