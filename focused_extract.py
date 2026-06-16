from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json

path = "/Users/zhangliang/Downloads/【排量】书法项目-排期&营收-规划表.xlsx"
wb = load_workbook(path, data_only=False, read_only=False)

def value(cell):
    v = cell.value
    if v is None:
        return ""
    if hasattr(v, "isoformat"):
        return v.isoformat()
    return v

result = {"sheets": []}
for ws in wb.worksheets:
    formula_cells = []
    non_empty_rows = []
    for row in ws.iter_rows():
        vals = [value(c) for c in row]
        if any(v != "" for v in vals):
            trimmed = vals[:]
            while trimmed and trimmed[-1] == "":
                trimmed.pop()
            non_empty_rows.append((row[0].row, trimmed[:30]))
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("=") and len(formula_cells) < 12:
                formula_cells.append({
                    "cell": f"{get_column_letter(c.column)}{c.row}",
                    "formula": c.value[:220],
                })

    result["sheets"].append({
        "sheet": ws.title,
        "size": [ws.max_row, ws.max_column],
        "first_non_empty_rows": non_empty_rows[:12],
        "sample_formula_cells": formula_cells,
    })

print(json.dumps(result, ensure_ascii=False, indent=2))
