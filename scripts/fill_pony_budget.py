from __future__ import annotations

import json
import math
import re
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = Path("/Users/zhangliang/Desktop/Pony表营期规划-转化率版本.xlsx")
CONFIG_PATH = ROOT / "data" / "config.json"
OUTPUT_DIR = ROOT / "exports"
MONTH = "2026-06"
MONTH_LABEL = f"{int(MONTH[-2:])}月"
OUTPUT_PATH = OUTPUT_DIR / f"Pony表营期规划-系统预算填充版-{MONTH}.xlsx"
PRODUCT_PRICE = 2580
SKU = "书画"

D_STAGES = [f"D{i}" for i in range(4, 14)]


def parse_day(value: str | None):
    if not value:
        return None
    return datetime.strptime(str(value)[:10], "%Y-%m-%d")


def add_days(day: str, count: int) -> str:
    return (parse_day(day) + timedelta(days=count)).strftime("%Y-%m-%d")


def days_between(start: str | None, end: str | None) -> list[str]:
    start_dt = parse_day(start)
    end_dt = parse_day(end)
    if not start_dt or not end_dt:
        return []
    days = []
    current = start_dt
    while current <= end_dt:
        days.append(current.strftime("%Y-%m-%d"))
        current += timedelta(days=1)
    return days


def parse_local_datetime(value: str):
    text = str(value or "")
    if text.endswith("T24:00"):
        return datetime.strptime(text[:10], "%Y-%m-%d") + timedelta(days=1)
    if "T" in text:
        return datetime.strptime(text[:16], "%Y-%m-%dT%H:%M")
    return datetime.strptime(text[:10], "%Y-%m-%d")


def target_weight_for_day(campaign: dict, day: str) -> float:
    if not campaign.get("intakeStartDateTime") or not campaign.get("intakeEndDateTime"):
        return 1.0 if campaign.get("intakeStart") <= day <= campaign.get("intakeEnd") else 0.0
    window_start = parse_local_datetime(campaign["intakeStartDateTime"])
    window_end = parse_local_datetime(campaign["intakeEndDateTime"])
    day_start = parse_local_datetime(f"{day}T00:00")
    day_end = parse_local_datetime(f"{day}T24:00")
    overlap = max(0.0, (min(window_end, day_end) - max(window_start, day_start)).total_seconds())
    return overlap / 86400


def campaign_target_share_map(campaigns: list[dict]) -> dict[str, float]:
    groups: dict[str, int] = {}
    for campaign in campaigns:
        key = campaign.get("openDate") or campaign.get("name")
        groups[key] = groups.get(key, 0) + 1
    return {
        campaign["name"]: 1 / max(groups.get(campaign.get("openDate") or campaign.get("name"), 1), 1)
        for campaign in campaigns
    }


def campaign_leads_by_subchannel(campaign: dict, targets: dict[tuple[str, str], float], share: float) -> dict[str, float]:
    result = {}
    for sub_id in campaign.get("subchannelIds", []):
        total = 0.0
        for day in days_between(campaign.get("intakeStart"), campaign.get("intakeEnd")):
            total += targets.get((day, sub_id), 0.0) * target_weight_for_day(campaign, day)
        result[sub_id] = total * share
    return result


def month_column_name(suffix: str) -> str:
    return f"{MONTH_LABEL}{suffix}"


def campaign_touches_month(campaign: dict, month: str) -> bool:
    intake_days = days_between(campaign.get("intakeStart"), campaign.get("intakeEnd"))
    if any(day.startswith(month) for day in intake_days):
        return True
    open_date = campaign.get("openDate")
    if not open_date:
        return False
    return any(add_days(open_date, int(stage[1:]) - 1).startswith(month) for stage in D_STAGES)


def mmdd_from_name(name: str) -> str:
    match = re.search(r"\.(\d{4})$", name or "")
    return match.group(1) if match else ""


def copy_row_style(ws, source_row: int, target_row: int):
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def main():
    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    campaigns = sorted(config.get("campaigns", []), key=lambda x: (x.get("openDate") or "", x.get("name") or ""))
    subchannels = config.get("subchannels", [])
    sub_by_id = {item["id"]: item for item in subchannels}
    sub_order = {item["id"]: index for index, item in enumerate(subchannels)}
    targets = {
        (item["date"], item["subchannelId"]): float(item.get("leads") or 0)
        for item in config.get("leadTargets", [])
    }
    r_values = {
        (item["subchannelId"], item["stage"]): float(item.get("rValue") or 0)
        for item in config.get("rTemplates", [])
    }
    for campaign in campaigns:
        for item in campaign.get("rOverrides", []):
            r_values[(item["subchannelId"], item["stage"])] = float(item.get("rValue") or 0)

    share_by_campaign = campaign_target_share_map(campaigns)
    rows = []
    for campaign in campaigns:
        if not campaign_touches_month(campaign, MONTH):
            continue
        leads_by_sub = campaign_leads_by_subchannel(campaign, targets, share_by_campaign.get(campaign["name"], 1))
        open_date = campaign.get("openDate") or ""
        conversion_start = add_days(open_date, 3) if open_date else ""
        close_date = add_days(open_date, 12) if open_date else ""
        conversion_dates = {stage: add_days(open_date, int(stage[1:]) - 1) for stage in D_STAGES} if open_date else {}
        for sub_id in campaign.get("subchannelIds", []):
            leads = leads_by_sub.get(sub_id, 0.0)
            if leads <= 0:
                continue
            sub = sub_by_id.get(sub_id, {"name": sub_id})
            daily_rates = [r_values.get((sub_id, stage), 0.0) / PRODUCT_PRICE for stage in D_STAGES]
            month_rates = [
                daily_rates[index]
                for index, stage in enumerate(D_STAGES)
                if conversion_dates.get(stage, "").startswith(MONTH)
            ]
            month_conversion_rate = sum(month_rates)
            rows.append({
                "_subOrder": sub_order.get(sub_id, 999),
                "_openDate": open_date,
                "渠道": sub["name"],
                "营期": campaign["name"],
                "sku": SKU,
                "转化产品单价": PRODUCT_PRICE,
                "接量时间": (campaign.get("intakeStartDateTime") or campaign.get("intakeStart") or "")[:10],
                "接量截止时间": (campaign.get("intakeEndDateTime") or campaign.get("intakeEnd") or "")[:10],
                "开课时间": open_date,
                "转化时间": conversion_start,
                "封板时间": close_date,
                "lead数量": round(leads),
                "leads单价": "",
                "转化率-汇总": sum(daily_rates),
                **{f"转化率-DAY{i + 1}": daily_rates[i] for i in range(10)},
                "当月转化率": month_conversion_rate,
                month_column_name("对应GMV"): round(leads * PRODUCT_PRICE * month_conversion_rate, 2),
                month_column_name("leads数量"): round(leads) if month_conversion_rate > 0 else 0,
                month_column_name("对应消耗"): "",
            })
    rows.sort(key=lambda row: (row["_subOrder"], row["_openDate"], row["营期"]))

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["Sheet1"]
    ws["X3"] = month_column_name("对应GMV")
    ws["Y3"] = month_column_name("leads数量")
    ws["Z3"] = month_column_name("对应消耗")
    headers = [ws.cell(3, col).value for col in range(1, ws.max_column + 1)]

    if ws.max_row > 3:
        ws.delete_rows(4, ws.max_row - 3)

    for index, row in enumerate(rows, start=4):
        copy_row_style(ws, 3, index)
        for col, header in enumerate(headers, start=1):
            if not header:
                continue
            ws.cell(index, col).value = row.get(header, "")
        for col in range(5, 10):
            ws.cell(index, col).number_format = "yyyy-mm-dd hh:mm"
        for col in range(12, 24):
            ws.cell(index, col).number_format = "0.00%"
        for col in range(24, 27):
            ws.cell(index, col).number_format = "#,##0"

    totals = {
        "lead数量": sum(row["lead数量"] for row in rows),
        month_column_name("对应GMV"): sum(row[month_column_name("对应GMV")] for row in rows),
        month_column_name("leads数量"): sum(row[month_column_name("leads数量")] for row in rows),
    }
    ws["A1"] = totals["lead数量"]
    ws["B1"] = totals[month_column_name("leads数量")]
    ws["C1"] = totals[month_column_name("对应GMV")]
    ws["D1"] = ""

    OUTPUT_DIR.mkdir(exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(json.dumps({
        "output": str(OUTPUT_PATH),
        "rows": len(rows),
        "totals": totals,
        "sample": rows[:5],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
