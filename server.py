from __future__ import annotations

import cgi
import csv
import json
import os
import re
import shutil
import socket
import subprocess
import tempfile
import urllib.error
import urllib.request
from copy import copy
from io import TextIOWrapper
from collections import defaultdict
from datetime import date, datetime, timedelta
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote, urlparse

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).parent
PUBLIC = ROOT / "public"
DATA_DIR = ROOT / "data"
CONFIG_PATH = DATA_DIR / "config.json"
OPERATION_LOG_PATH = DATA_DIR / "operation-log.jsonl"
EXPORT_DIR = ROOT / "exports"
PONY_TEMPLATE_PATH = ROOT / "templates" / "Pony表营期规划-转化率版本.xlsx"
FEISHU_NOTIFY_PATH = ROOT / ".local" / "feishu-notify.json"
REVENUE_TARGET_CACHE_PATH = DATA_DIR / "revenue-targets.json"
REVENUE_ACTUALS_CACHE_PATH = DATA_DIR / "revenue-actuals.json"
CRM_NORMALIZED_DIR = DATA_DIR / "integration-checks" / "normalized"
CRM_STORAGE_STATE_PATH = ROOT / ".local" / "crm-storage-state.json"
CRM_REFRESH_SCRIPT_PATH = ROOT / "scripts" / "refresh_crm_login.mjs"
FEISHU_REPORT_SCRIPT_PATH = ROOT / "scripts" / "send_revenue_overview_card.mjs"
FALLBACK_NODE_PATH = Path.home() / ".cache/codex-runtimes/codex-primary-runtime/dependencies/node/bin/node"
CRM_KEYCHAIN_SERVICE = "crm-dashboard"
CRM_KEYCHAIN_ACCOUNT = "zhangliang0102"
CRM_ORDER_LIST_URL = "https://kapi.likeduoduiyi.cn/kk/cms/order/list"
D_STAGES = [f"D{i}" for i in range(4, 14)]
PONY_PRODUCT_PRICE = 2580
REVENUE_SPREADSHEET_TOKEN = "SVa6s6c31hztbRt0XE0cjxkCnih"
REVENUE_SHEET_ID = "90a237"
REVENUE_SEGMENTS = [
    {"business": "书法", "side": "前端", "label": "书法前端", "targetRow": 4, "orderKind": "销转订单", "crmBusiness": "书法", "category": 3, "type": "1", "businessDepartId": 1},
    {"business": "书法", "side": "后端", "label": "书法后端", "targetRow": 6, "orderKind": "学管扩转续订单", "crmBusiness": "书法", "category": 3, "type": "2", "businessDepartId": 2},
    {"business": "朗诵", "side": "后端", "label": "朗诵后端", "targetRow": 8, "orderKind": "学管扩转续订单", "crmBusiness": "朗诵", "category": 14, "type": "2", "businessDepartId": 2},
]
FEISHU_REPORT_SCOPES = {"overall", "calligraphy_front", "calligraphy_backend", "recitation_backend"}
SERVER_START_TIME = datetime.now().isoformat(timespec="seconds")


def local_lan_ip() -> str:
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
            sock.connect(("8.8.8.8", 80))
            return sock.getsockname()[0]
    except OSError:
        return "127.0.0.1"


def default_config() -> dict[str, Any]:
    return {
        "projectName": "书法",
        "channels": [
            {"id": "bd1", "name": "BD1"},
            {"id": "paid", "name": "信息流"},
            {"id": "private", "name": "私域"},
            {"id": "other", "name": "其他"},
        ],
        "subchannels": [
            {"id": "bd1_books", "channelId": "bd1", "name": "BD1-图书"},
            {"id": "bd1_free", "channelId": "bd1", "name": "BD1-0元"},
            {"id": "paid_mix", "channelId": "paid", "name": "信息流-混合"},
            {"id": "miniapp", "channelId": "private", "name": "小程序"},
            {"id": "wecom", "channelId": "private", "name": "企微析出"},
        ],
        "teachers": [
            {"code": "BZ", "name": "白止"},
            {"code": "ST", "name": "ST"},
        ],
        "leadTargets": [],
        "rTemplates": [],
        "campaigns": [],
        "actualCampaigns": [],
        "studioMappings": DEFAULT_TEAM_MAPPINGS,
        "intakeRules": [
            {
                "id": "standard_3_5",
                "name": "标准3.5天切量",
                "allocation": "hourly",
                "isDefault": True,
                "entries": [
                    {"openWeekday": 1, "startWeekday": 3, "startTime": "10:00", "endWeekday": 6, "endTime": "22:00"},
                    {"openWeekday": 4, "startWeekday": 6, "startTime": "22:00", "endWeekday": 3, "endTime": "10:00"},
                ],
            }
        ],
        "budgetSnapshots": [],
        "predictionSnapshots": [],
        "settings": {
            "gmvDanger": 80,
            "leadsLow": 80,
            "leadsHigh": 120,
            "leadsGoodLow": 90,
            "leadsGoodHigh": 110,
        },
    }


def read_config() -> dict[str, Any]:
    DATA_DIR.mkdir(exist_ok=True)
    if not CONFIG_PATH.exists():
        config = default_config()
        write_config(config)
        return config
    with CONFIG_PATH.open("r", encoding="utf-8") as handle:
        config = json.load(handle)
    merged = default_config()
    merged.update(config)
    for key, value in default_config().items():
        if key not in merged:
            merged[key] = value
    return merged


def write_config(config: dict[str, Any]) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    if CONFIG_PATH.exists():
        backup_dir = DATA_DIR / "backups"
        backup_dir.mkdir(exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup_path = backup_dir / f"config-{stamp}.json"
        backup_path.write_text(CONFIG_PATH.read_text(encoding="utf-8"), encoding="utf-8")
    with CONFIG_PATH.open("w", encoding="utf-8") as handle:
        json.dump(config, handle, ensure_ascii=False, indent=2)


def append_operation_log(entry: dict[str, Any]) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    payload = {
        "serverTime": datetime.now().isoformat(timespec="seconds"),
        **entry,
    }
    with OPERATION_LOG_PATH.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(payload, ensure_ascii=False) + "\n")


def read_operation_log(limit: int = 200) -> list[dict[str, Any]]:
    if not OPERATION_LOG_PATH.exists():
        return []
    lines = OPERATION_LOG_PATH.read_text(encoding="utf-8").splitlines()[-limit:]
    result = []
    for line in lines:
        try:
            result.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return result


def http_json(url: str, method: str = "GET", payload: dict[str, Any] | None = None, token: str = "") -> dict[str, Any]:
    body = None
    headers = {"Content-Type": "application/json; charset=utf-8"}
    if payload is not None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    if token:
        headers["Authorization"] = f"Bearer {token}"
    request = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        text = exc.read().decode("utf-8", errors="replace")
        try:
            payload = json.loads(text)
        except json.JSONDecodeError:
            payload = {"msg": text}
        payload["_httpStatus"] = exc.code
        return payload


def feishu_tenant_access_token() -> str:
    if not FEISHU_NOTIFY_PATH.exists():
        raise RuntimeError("缺少 .local/feishu-notify.json，无法读取飞书目标表。")
    config = json.loads(FEISHU_NOTIFY_PATH.read_text(encoding="utf-8"))
    if not config.get("app_id") or not config.get("app_secret"):
        raise RuntimeError("飞书应用配置缺少 app_id 或 app_secret。")
    payload = http_json(
        "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
        method="POST",
        payload={"app_id": config["app_id"], "app_secret": config["app_secret"]},
    )
    if payload.get("code") != 0:
        raise RuntimeError(f"获取飞书 tenant_access_token 失败：{payload.get('code')} {payload.get('msg')}")
    return str(payload["tenant_access_token"])


def read_feishu_notify_config() -> dict[str, Any]:
    if not FEISHU_NOTIFY_PATH.exists():
        raise RuntimeError("缺少 .local/feishu-notify.json，无法发送飞书日报。")
    return json.loads(FEISHU_NOTIFY_PATH.read_text(encoding="utf-8"))


def feishu_report_targets() -> list[dict[str, str]]:
    config = read_feishu_notify_config()
    targets = config.get("targets")
    if isinstance(targets, list) and targets:
        safe_targets = []
        for item in targets:
            if not isinstance(item, dict) or not item.get("id") or not item.get("receive_id"):
                continue
            safe_targets.append({
                "id": str(item["id"]),
                "label": str(item.get("label") or item["id"]),
            })
        return safe_targets
    if config.get("receive_id"):
        return [{"id": "default", "label": str(config.get("label") or "默认接收群")}]
    return []


def send_feishu_overview_report(month: str, scope: str, target_id: str) -> dict[str, Any]:
    month = str(month or "")
    if not re.match(r"^\d{4}-\d{2}$", month):
        month = datetime.now().strftime("%Y-%m")
    if scope not in FEISHU_REPORT_SCOPES:
        raise RuntimeError(f"不支持的日报范围：{scope}")
    targets = feishu_report_targets()
    if not targets:
        raise RuntimeError("飞书配置里没有可发送的接收群。")
    if target_id and not any(item["id"] == target_id for item in targets):
        raise RuntimeError(f"没有找到飞书接收群：{target_id}")
    if not target_id:
        target_id = targets[0]["id"]
    base_url = f"http://127.0.0.1:{int(os.environ.get('PORT', '8765'))}"
    node_binary = shutil.which("node") or (str(FALLBACK_NODE_PATH) if FALLBACK_NODE_PATH.exists() else "node")
    command = [
        node_binary,
        str(FEISHU_REPORT_SCRIPT_PATH),
        "--month",
        month,
        "--scope",
        scope,
        "--target-id",
        target_id,
        "--base-url",
        base_url,
    ]
    result = subprocess.run(command, cwd=ROOT, capture_output=True, text=True, timeout=60, check=False)
    if result.returncode != 0:
        detail = (result.stderr or result.stdout or "").strip()
        raise RuntimeError(detail or "飞书日报发送失败。")
    try:
        payload = json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"飞书日报发送结果解析失败：{exc}") from exc
    return payload


def feishu_callback_card(value: dict[str, Any]) -> dict[str, Any]:
    tabs = value.get("tabs") if isinstance(value.get("tabs"), list) else []
    selected_scope = str(value.get("selectedScope") or "overall")
    selected = next((item for item in tabs if isinstance(item, dict) and item.get("scope") == selected_scope), None)
    if selected is None and tabs:
        selected = tabs[0] if isinstance(tabs[0], dict) else {}
    selected = selected or {}
    title = str(value.get("title") or "经营进度日报 · 书法（Demo）")
    elements: list[dict[str, Any]] = [
        {
            "tag": "markdown",
            "content": str(selected.get("summaryText") or "**书法｜营收目标口径**"),
        }
    ]
    if selected.get("imageKey"):
        elements.append({
            "tag": "img",
            "img_key": str(selected["imageKey"]),
            "alt": {"tag": "plain_text", "content": f"{selected.get('label') or '经营进度'}看板"},
        })
    elements.extend([
        {"tag": "hr"},
        {
            "tag": "markdown",
            "content": "\n".join([
                "**口径说明**",
                "· 数据源：本地看板 CRM 营收总览",
                "· 实际：CRM totalPriceString，按 createTime 归日",
                "· 点击底部按钮可切换不同业务视图",
            ]),
        },
        {
            "tag": "action",
            "actions": [
                {
                    "tag": "button",
                    "text": {"tag": "plain_text", "content": str(tab.get("label") or tab.get("scope") or "视图")},
                    "type": "primary" if tab.get("scope") == selected_scope else "default",
                    "value": {
                        "mode": "revenue_overview_demo",
                        "month": value.get("month"),
                        "title": title,
                        "selectedScope": tab.get("scope"),
                        "tabs": tabs,
                    },
                }
                for tab in tabs
                if isinstance(tab, dict)
            ],
        },
    ])
    return {
        "config": {"wide_screen_mode": True},
        "header": {
            "template": "blue",
            "title": {"tag": "plain_text", "content": title},
        },
        "elements": elements,
    }


def extract_feishu_card_value(payload: dict[str, Any]) -> dict[str, Any]:
    candidates = [
        payload.get("event", {}).get("action", {}).get("value") if isinstance(payload.get("event"), dict) else None,
        payload.get("action", {}).get("value") if isinstance(payload.get("action"), dict) else None,
        payload.get("value"),
    ]
    for candidate in candidates:
        if isinstance(candidate, dict):
            return candidate
        if isinstance(candidate, str):
            try:
                parsed = json.loads(candidate)
            except json.JSONDecodeError:
                continue
            if isinstance(parsed, dict):
                return parsed
    return {}


def handle_feishu_card_callback(payload: dict[str, Any]) -> dict[str, Any]:
    if payload.get("challenge"):
        return {"challenge": payload["challenge"]}
    if payload.get("type") == "url_verification" and payload.get("challenge"):
        return {"challenge": payload["challenge"]}
    config = read_feishu_notify_config()
    expected_token = config.get("verification_token")
    actual_token = None
    if isinstance(payload.get("header"), dict):
        actual_token = payload["header"].get("token")
    actual_token = actual_token or payload.get("token")
    if expected_token and actual_token and str(expected_token) != str(actual_token):
        raise RuntimeError("飞书回调 token 校验失败。")
    value = extract_feishu_card_value(payload)
    if value.get("mode") != "revenue_overview_demo":
        return {"toast": {"type": "info", "content": "已收到卡片操作。"}}
    selected_label = next((str(tab.get("label")) for tab in value.get("tabs", []) if isinstance(tab, dict) and tab.get("scope") == value.get("selectedScope")), "对应视图")
    return {
        "toast": {"type": "success", "content": f"已切换到{selected_label}"},
        "card": {
            "type": "raw",
            "data": feishu_callback_card(value),
        },
    }


def excel_serial_to_day(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, str) and re.match(r"\d{4}-\d{1,2}-\d{1,2}", value):
        return normalize_date(value) or ""
    try:
        serial = float(value)
    except (TypeError, ValueError):
        return ""
    return (date(1899, 12, 30) + timedelta(days=int(serial))).isoformat()


def revenue_target_cache() -> dict[str, Any]:
    if not REVENUE_TARGET_CACHE_PATH.exists():
        return {"records": [], "source": {"cached": False}}
    try:
        return json.loads(REVENUE_TARGET_CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"records": [], "source": {"cached": False, "error": "本地营收目标缓存无法解析"}}


def revenue_actuals_cache() -> dict[str, Any]:
    if not REVENUE_ACTUALS_CACHE_PATH.exists():
        return {"records": [], "source": [], "error": "本地 CRM 实际缓存不存在"}
    try:
        return json.loads(REVENUE_ACTUALS_CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"records": [], "source": [], "error": "本地 CRM 实际缓存无法解析"}


def local_stamp() -> str:
    return datetime.now().strftime("%Y%m%d%H%M%S")


def crm_sync_days(month: str, end_date: str = "") -> list[str]:
    year, month_index = [int(part) for part in month.split("-")]
    next_month = date(year + int(month_index == 12), 1 if month_index == 12 else month_index + 1, 1)
    month_end = next_month - timedelta(days=1)
    today = date.today()
    final_day = month_end
    if end_date and end_date.startswith(month):
        final_day = min(datetime.strptime(end_date, "%Y-%m-%d").date(), month_end)
    elif today.strftime("%Y-%m") == month:
        final_day = min(today, month_end)
    return [date(year, month_index, day).isoformat() for day in range(1, final_day.day + 1)]


class CrmLoginExpired(RuntimeError):
    pass


def is_crm_login_expired(status: Any, message: Any) -> bool:
    text = str(message or "")
    return str(status) == "2000" or any(keyword in text for keyword in ["重新登录", "其他设备登录", "登录态", "token"])


def node_binary() -> str:
    candidates = [
        os.environ.get("NODE_BINARY", ""),
        "/Users/zhangliang/.cache/codex-runtimes/codex-primary-runtime/dependencies/node/bin/node",
        shutil.which("node") or "",
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return candidate
    raise RuntimeError("找不到 Node.js，无法自动刷新 CRM 登录。")


def crm_refresh_login_from_keychain() -> None:
    if not CRM_REFRESH_SCRIPT_PATH.exists():
        raise RuntimeError("找不到 CRM 登录刷新脚本。")
    env = os.environ.copy()
    node_modules = ROOT / "node_modules"
    bundled_node_modules = Path("/Users/zhangliang/.cache/codex-runtimes/codex-primary-runtime/dependencies/node/node_modules")
    node_path_items = [str(path) for path in [node_modules, bundled_node_modules] if path.exists()]
    if node_path_items:
        env["NODE_PATH"] = os.pathsep.join(node_path_items + ([env["NODE_PATH"]] if env.get("NODE_PATH") else []))
    command = [
        node_binary(),
        str(CRM_REFRESH_SCRIPT_PATH),
        "--auto-keychain",
        "--headless",
        "--timeout-ms",
        "120000",
        "--service",
        CRM_KEYCHAIN_SERVICE,
        "--account",
        CRM_KEYCHAIN_ACCOUNT,
    ]
    try:
        result = subprocess.run(
            command,
            cwd=ROOT,
            env=env,
            check=False,
            capture_output=True,
            text=True,
            timeout=150,
        )
    except subprocess.TimeoutExpired as exc:
        raise RuntimeError("CRM 自动登录超时，旧数据已保留。") from exc
    if result.returncode != 0:
        detail = "\n".join(part.strip() for part in [result.stderr, result.stdout] if part and part.strip())
        if "find-generic-password" in detail or "钥匙串" in detail:
            raise RuntimeError("CRM 自动登录失败：无法读取钥匙串凭据，请检查 crm-dashboard 是否已保存。")
        if "登录输入框" in detail or "登录按钮" in detail:
            raise RuntimeError("CRM 自动登录失败：登录页结构可能变化，请手动登录后再试。")
        raise RuntimeError("CRM 自动登录失败，旧数据已保留。")


def crm_token_from_storage() -> str:
    if not CRM_STORAGE_STATE_PATH.exists():
        raise CrmLoginExpired("本地 CRM 登录态不存在，需要刷新登录。")
    state = json.loads(CRM_STORAGE_STATE_PATH.read_text(encoding="utf-8"))
    origins = state.get("origins", [])
    storage = next((item.get("localStorage", []) for item in origins if item.get("origin") == "https://kkhc-admin.likeduoduiyi.cn"), [])
    admin_info_raw = next((item.get("value", "") for item in storage if item.get("name") == "admin_info"), "")
    admin_info = json.loads(admin_info_raw or "{}")
    token = admin_info.get("token")
    if not token:
        raise CrmLoginExpired("CRM 登录态里没有 token，需要刷新登录。")
    return str(token)


def crm_order_payload(day: str, segment: dict[str, Any]) -> dict[str, Any]:
    return {
        "category": segment["category"],
        "goodsId": "",
        "nickName": "",
        "orderTime": [f"{day} 00:00:00", f"{day} 23:59:59"],
        "payTime": [],
        "payStatus": "",
        "payType": "",
        "campIds": [],
        "empIds": [],
        "orderNo": "",
        "inClass": "",
        "type": segment["type"],
        "addAst": "",
        "classCampId": "",
        "outNo": "",
        "needJudge": "",
        "isCombine": "",
        "isUp": "",
        "empNum": "",
        "frontEnd": "",
        "isSale": "",
        "saleCampId": "",
        "astId": "",
        "handoverCampId": "",
        "refundStatus": "",
        "businessDepartId": segment["businessDepartId"],
        "invoiceStatus": "",
        "redFlag": "",
        "isHaveAddress": "",
        "kkTeamId": "",
        "kkGroupId": "",
        "auditStatus": "",
        "unionId": "",
        "startTime": f"{day} 00:00:00",
        "endTime": f"{day} 23:59:59",
        "current": 1,
        "size": 20,
    }


def crm_fetch_segment_day(token: str, day: str, segment: dict[str, Any]) -> dict[str, Any]:
    payload = crm_order_payload(day, segment)
    body = json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(
        CRM_ORDER_LIST_URL,
        data=body,
        headers={
            "accept": "application/json, text/plain, */*",
            "content-type": "application/json;charset=UTF-8",
            "referer": "https://kkhc-admin.likeduoduiyi.cn/",
            "token": token,
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            response_payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        message = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"{segment['label']} {day} CRM请求失败：HTTP {exc.code} {message}") from exc
    status = response_payload.get("status")
    message = response_payload.get("message", "")
    if is_crm_login_expired(status, message):
        raise CrmLoginExpired(f"{segment['label']} {day} CRM登录已失效：{status} {message}")
    if response_payload.get("status") != 200:
        raise RuntimeError(f"{segment['label']} {day} CRM请求失败：{status} {message}")
    data = response_payload.get("data") or {}
    return {
        "date": day,
        "business": segment["business"],
        "side": segment["side"],
        "label": segment["label"],
        "amount": to_number(data.get("totalPriceString")),
        "orderKind": segment["orderKind"],
        "crmBusiness": segment["crmBusiness"],
        "rowCount": int(to_number(data.get("total"))),
        "source": "crm",
        "sourceField": "totalPriceString",
        "timeField": "createTime",
    }


def sync_revenue_actuals_from_crm(month: str, end_date: str = "") -> dict[str, Any]:
    if not re.match(r"^\d{4}-\d{2}$", month or ""):
        raise RuntimeError("月份格式不正确，应为 YYYY-MM。")
    days = crm_sync_days(month, end_date)
    if not days:
        raise RuntimeError("没有可同步的日期。")
    login_refreshed = False

    def fetch_records() -> list[dict[str, Any]]:
        token = crm_token_from_storage()
        return [
            crm_fetch_segment_day(token, day, segment)
            for day in days
            for segment in REVENUE_SEGMENTS
        ]

    try:
        records = fetch_records()
    except CrmLoginExpired:
        crm_refresh_login_from_keychain()
        login_refreshed = True
        records = fetch_records()
    source = []
    for segment in REVENUE_SEGMENTS:
        segment_records = [item for item in records if item["label"] == segment["label"]]
        source.append({
            "label": segment["label"],
            "business": segment["business"],
            "side": segment["side"],
            "orderKind": segment["orderKind"],
            "crmBusiness": segment["crmBusiness"],
            "totalAmount": sum(to_number(item.get("amount")) for item in segment_records),
            "rowCount": sum(to_number(item.get("rowCount")) for item in segment_records),
        })
    output = {
        "syncedAt": datetime.now().isoformat(timespec="seconds"),
        "range": {
            "start": days[0],
            "end": days[-1],
            "timeField": "createTime",
            "payTime": "empty",
            "sourceField": "totalPriceString",
        },
        "records": records,
        "source": source,
    }
    backup_path = ""
    backup_dir = DATA_DIR / "backups"
    if REVENUE_ACTUALS_CACHE_PATH.exists():
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = str(backup_dir / f"revenue-actuals-{local_stamp()}.json")
        shutil.copyfile(REVENUE_ACTUALS_CACHE_PATH, backup_path)
    DATA_DIR.mkdir(exist_ok=True)
    REVENUE_ACTUALS_CACHE_PATH.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    report_dir = DATA_DIR / "integration-checks" / "reports"
    report_dir.mkdir(parents=True, exist_ok=True)
    report_path = report_dir / f"crm-revenue-sync-{month}-{local_stamp()}.json"
    report_path.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "ok": True,
        "month": month,
        "range": output["range"],
        "source": source,
        "syncedAt": output["syncedAt"],
        "recordCount": len(records),
        "loginRefreshed": login_refreshed,
        "reportPath": str(report_path),
        "backupPath": backup_path,
    }


def sync_revenue_targets_from_feishu() -> dict[str, Any]:
    token = feishu_tenant_access_token()
    sheets = http_json(
        f"https://open.feishu.cn/open-apis/sheets/v3/spreadsheets/{REVENUE_SPREADSHEET_TOKEN}/sheets/query",
        token=token,
    )
    if sheets.get("code") != 0:
        raise RuntimeError(f"读取飞书工作表失败：{sheets.get('code')} {sheets.get('msg')}")
    range_id = quote(f"{REVENUE_SHEET_ID}!A1:AF20", safe="")
    values_payload = http_json(
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{REVENUE_SPREADSHEET_TOKEN}/values/{range_id}",
        token=token,
    )
    if values_payload.get("code") != 0:
        raise RuntimeError(f"读取飞书目标范围失败：{values_payload.get('code')} {values_payload.get('msg')}")
    values = values_payload.get("data", {}).get("valueRange", {}).get("values", [])
    if len(values) < 9:
        raise RuntimeError("飞书目标表行数不足，无法读取第 4/5/7/9 行。")
    date_row = values[3]
    records: list[dict[str, Any]] = []
    for segment in REVENUE_SEGMENTS:
        row = values[segment["targetRow"]]
        for col in range(2, max(len(date_row), len(row))):
            day = excel_serial_to_day(date_row[col] if col < len(date_row) else "")
            if not day:
                continue
            target = to_number(row[col] if col < len(row) else 0)
            records.append({
                "date": day,
                "business": segment["business"],
                "side": segment["side"],
                "label": segment["label"],
                "targetAmount": target,
                "source": "feishu",
            })
    sheet_title = next((item.get("title") for item in sheets.get("data", {}).get("sheets", []) if item.get("sheet_id") == REVENUE_SHEET_ID), "统计")
    cache = {
        "syncedAt": datetime.now().isoformat(timespec="seconds"),
        "records": records,
        "source": {
            "cached": False,
            "spreadsheetToken": REVENUE_SPREADSHEET_TOKEN,
            "sheetId": REVENUE_SHEET_ID,
            "sheetTitle": sheet_title,
            "range": f"{REVENUE_SHEET_ID}!A1:AF20",
        },
    }
    DATA_DIR.mkdir(exist_ok=True)
    REVENUE_TARGET_CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    return cache


def latest_crm_file(month: str, order_kind: str, business: str) -> Path | None:
    if not CRM_NORMALIZED_DIR.exists():
        return None
    pattern = f"crm-{month}-{order_kind}-{business}-*.json"
    files = sorted(CRM_NORMALIZED_DIR.glob(pattern))
    return files[-1] if files else None


def read_revenue_actuals(month: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    cache = revenue_actuals_cache()
    records = [
        item for item in cache.get("records", [])
        if str(item.get("date", "")).startswith(month)
    ]
    sources = cache.get("source", [])
    return records, sources


def week_start_day(day: str) -> str:
    parsed = datetime.strptime(day, "%Y-%m-%d").date()
    return (parsed - timedelta(days=parsed.weekday())).isoformat()


def revenue_bucket(target: float = 0, actual: float = 0) -> dict[str, Any]:
    diff = actual - target
    return {
        "targetAmount": target,
        "actualAmount": actual,
        "achievementRate": actual / target if target else None,
        "diffAmount": diff,
        "hasTarget": target > 0,
        "targetlessActual": target <= 0 and actual > 0,
    }


def revenue_overview(month: str) -> dict[str, Any]:
    warnings: list[str] = []
    target_cache = revenue_target_cache()
    if not target_cache.get("records"):
        warnings.append("本地定版目标为空，请先完成飞书目标校验并锁定。")
    actual_cache = revenue_actuals_cache()
    if actual_cache.get("error"):
        warnings.append(str(actual_cache["error"]))
    targets = [item for item in target_cache.get("records", []) if str(item.get("date", "")).startswith(month)]
    actuals, actual_sources = read_revenue_actuals(month)
    days = sorted({item["date"] for item in targets} | {item["date"] for item in actuals})
    labels = [item["label"] for item in REVENUE_SEGMENTS]
    total_target = sum(to_number(item.get("targetAmount")) for item in targets)
    total_actual = sum(to_number(item.get("amount")) for item in actuals)

    daily = []
    for day in days:
        target = sum(to_number(item.get("targetAmount")) for item in targets if item.get("date") == day)
        actual = sum(to_number(item.get("amount")) for item in actuals if item.get("date") == day)
        daily.append({"date": day, **revenue_bucket(target, actual)})

    weekly_map: dict[str, dict[str, float]] = defaultdict(lambda: {"target": 0, "actual": 0})
    for item in targets:
        weekly_map[week_start_day(item["date"])]["target"] += to_number(item.get("targetAmount"))
    for item in actuals:
        weekly_map[week_start_day(item["date"])]["actual"] += to_number(item.get("amount"))
    weekly = [
        {"weekStart": key, "weekEnd": (datetime.strptime(key, "%Y-%m-%d").date() + timedelta(days=6)).isoformat(), **revenue_bucket(value["target"], value["actual"])}
        for key, value in sorted(weekly_map.items())
    ]

    business_rows = []
    for segment in REVENUE_SEGMENTS:
        target = sum(to_number(item.get("targetAmount")) for item in targets if item.get("business") == segment["business"] and item.get("side") == segment["side"])
        actual = sum(to_number(item.get("amount")) for item in actuals if item.get("business") == segment["business"] and item.get("side") == segment["side"])
        business_rows.append({
            "business": segment["business"],
            "side": segment["side"],
            "label": segment["label"],
            **revenue_bucket(target, actual),
        })

    return {
        "month": month,
        "summary": revenue_bucket(total_target, total_actual),
        "daily": daily,
        "weekly": weekly,
        "business": business_rows,
        "targetRecords": targets,
        "actualRecords": actuals,
        "actualSources": actual_sources,
        "targetSource": target_cache.get("source", {}),
        "targetSyncedAt": target_cache.get("syncedAt", ""),
        "actualSyncedAt": actual_cache.get("syncedAt", ""),
        "actualRange": actual_cache.get("range", {}),
        "actualRecordCount": len(actuals),
        "targetRecordCount": len(targets),
        "businessLabels": labels,
        "warnings": warnings,
    }


def normalize_actual_campaign_row(row: dict[str, Any]) -> dict[str, Any] | None:
    name = row.get("name") or row.get("campaignName") or row.get("营期") or row.get("营期名") or row.get("营期名称")
    if not name:
        return None
    return {
        "name": str(name).strip(),
        "actualLeads": to_number(row.get("actualLeads") or row.get("leads") or row.get("leads数") or row.get("Leads") or row.get("实际Leads") or row.get("实际leads")),
        "actualGmv": to_number(row.get("actualGmv") or row.get("gmv") or row.get("GMV") or row.get("实际GMV") or row.get("实际gmv")),
        "spend": to_number(row.get("spend") or row.get("消耗")),
        "fullPriceStudents": to_number(row.get("fullPriceStudents") or row.get("正价课学员数") or row.get("正价学员数")),
        "categories": row.get("categories") or row.get("分类") or "",
        "openDate": normalize_date(row.get("openDate") or row.get("开课日期")) or "",
        "closeDate": normalize_date(row.get("closeDate") or row.get("封板日期")) or "",
        "stage": row.get("stage") or row.get("营期阶段") or "",
        "rValue": to_number(row.get("rValue") or row.get("累计R值")),
        "roi": to_number(row.get("roi") or row.get("ROI")),
        "conversionRate": to_number(row.get("conversionRate") or row.get("转化率")),
        "actualSubchannels": row.get("actualSubchannels") or [],
        "studioTotals": row.get("studioTotals") or [],
        "rBreakdown": row.get("rBreakdown") or r_breakdown_from_row(row),
        "source": row.get("source") or "导入",
    }


def r_breakdown_from_row(row: dict[str, Any]) -> dict[str, float]:
    result: dict[str, float] = {}
    for index in range(4, 15):
        stage = f"D{index}"
        value = to_number(row.get(f"d{index}_R值") or row.get(f"D{index}_R值") or row.get(f"D{index}-R值"))
        if value:
            result[f"{stage}-R值"] = value
    return result


def normalize_actual_subchannel(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "studio": row.get("工作室") or "",
        "category": row.get("分类") or row.get("categories") or "",
        "actualLeads": to_number(row.get("leads数") or row.get("actualLeads") or row.get("leads")),
        "actualGmv": to_number(row.get("GMV") or row.get("gmv") or row.get("actualGmv")),
        "spend": to_number(row.get("消耗") or row.get("spend")),
        "rValue": to_number(row.get("累计R值") or row.get("rValue")),
        "roi": to_number(row.get("ROI") or row.get("roi")),
        "conversionRate": to_number(row.get("转化率") or row.get("conversionRate")),
        "rBreakdown": r_breakdown_from_row(row),
    }


def group_native_actual_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not any("工作室" in row and "分类" in row and "营期" in row for row in rows):
        return rows
    groups: dict[str, dict[str, Any]] = {}
    for row in rows:
        name = str(row.get("营期") or "").strip()
        if not name:
            continue
        group = groups.setdefault(name, {"aggregate": None, "subchannels": [], "studioTotals": []})
        studio = str(row.get("工作室") or "").strip()
        category = str(row.get("分类") or "").strip()
        if studio == "营期汇总" and category == "-":
            group["aggregate"] = row
        elif category == "工作室汇总":
            group["studioTotals"].append(normalize_actual_subchannel(row))
        else:
            group["subchannels"].append(normalize_actual_subchannel(row))
    result: list[dict[str, Any]] = []
    for name, group in groups.items():
        aggregate = dict(group["aggregate"] or {"营期": name})
        aggregate["name"] = name
        aggregate["actualSubchannels"] = group["subchannels"]
        aggregate["studioTotals"] = group["studioTotals"]
        aggregate["categories"] = "、".join(item["category"] for item in group["subchannels"] if item.get("category"))
        aggregate["rBreakdown"] = r_breakdown_from_row(aggregate)
        aggregate["source"] = "营期渠道数据统计"
        result.append(aggregate)
    return result


def parse_actual_import(path: str, filename: str) -> list[dict[str, Any]]:
    lower = filename.lower()
    rows: list[dict[str, Any]] = []
    with open(path, "rb") as probe:
        prefix = probe.read(32).lstrip()
    if prefix.startswith(b"{") or prefix.startswith(b"["):
        with open(path, "r", encoding="utf-8-sig") as handle:
            payload = json.load(handle)
        if isinstance(payload, dict) and ("via" in payload or "trace" in payload):
            message = ""
            via = payload.get("via") or []
            if via and isinstance(via, list):
                message = via[0].get("message", "")
            raise ValueError(f"这个文件不是有效数据文件，而是导出系统返回的错误信息：{message or '未知错误'}")
        json_rows = payload if isinstance(payload, list) else payload.get("campaigns") or payload.get("actualCampaigns") or []
        json_rows = group_native_actual_rows(json_rows)
        return [item for item in (normalize_actual_campaign_row(row) for row in json_rows) if item]
    if lower.endswith(".csv"):
        with open(path, "rb") as handle:
            text = TextIOWrapper(handle, encoding="utf-8-sig", newline="")
            rows = list(csv.DictReader(text))
    elif lower.endswith((".xlsx", ".xlsm", ".xls")):
        workbook = load_workbook(path, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        values = list(sheet.iter_rows(values_only=True))
        if values:
            headers = [str(value or "").strip() for value in values[0]]
            for values_row in values[1:]:
                rows.append({headers[index]: value for index, value in enumerate(values_row) if index < len(headers)})
    else:
        raise ValueError("暂不支持该文件格式，请使用 CSV 或 XLSX。")
    rows = group_native_actual_rows(rows)
    return [item for item in (normalize_actual_campaign_row(row) for row in rows) if item]


def parse_iso_day(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def daterange(start: str | None, end: str | None) -> list[str]:
    start_day = parse_iso_day(start)
    end_day = parse_iso_day(end)
    if not start_day or not end_day or end_day < start_day:
        return []
    days = []
    current = start_day
    while current <= end_day:
        days.append(current.isoformat())
        current += timedelta(days=1)
    return days


def stage_for_day(open_date: str | None, current_date: str | None) -> str:
    open_day = parse_iso_day(open_date)
    current_day = parse_iso_day(current_date)
    if not open_day or not current_day:
        return ""
    delta = (current_day - open_day).days
    if delta >= 0:
        return f"D{delta + 1}"
    return "D0" if delta == -1 else f"D{delta + 1}"


def make_campaign_name(config: dict[str, Any], base_no: int, sub_no: int, teacher_code: str, open_date: str) -> str:
    day = parse_iso_day(open_date)
    suffix = day.strftime("%m%d") if day else str(open_date).replace("-", "")[-4:]
    return f"{config.get('projectName', '书法')}{base_no}.{sub_no}.{teacher_code}.{suffix}"


def target_lookup(config: dict[str, Any]) -> dict[tuple[str, str], float]:
    result: dict[tuple[str, str], float] = {}
    for item in config.get("leadTargets", []):
        day = normalize_date(item.get("date"))
        sub_id = str(item.get("subchannelId") or "")
        if day and sub_id:
            result[(day, sub_id)] = to_number(item.get("leads"))
    return result


def r_lookup(config: dict[str, Any], campaign: dict[str, Any] | None = None) -> dict[tuple[str, str], float]:
    result: dict[tuple[str, str], float] = {}
    for item in config.get("rTemplates", []):
        sub_id = str(item.get("subchannelId") or "")
        stage = str(item.get("stage") or "")
        if sub_id and stage:
            result[(sub_id, stage)] = to_number(item.get("rValue"))
    if campaign:
        for item in campaign.get("rOverrides", []):
            sub_id = str(item.get("subchannelId") or "")
            stage = str(item.get("stage") or "")
            if sub_id and stage:
                result[(sub_id, stage)] = to_number(item.get("rValue"))
    return result


def parse_local_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    text = str(value)
    try:
        if text.endswith("T24:00"):
            return datetime.strptime(text[:10], "%Y-%m-%d") + timedelta(days=1)
        if "T" in text:
            return datetime.strptime(text[:16], "%Y-%m-%dT%H:%M")
        return datetime.strptime(text[:10], "%Y-%m-%d")
    except ValueError:
        return None


def add_days(day: str | None, count: int) -> str:
    parsed = parse_iso_day(day)
    return (parsed + timedelta(days=count)).isoformat() if parsed else ""


def target_weight_for_day(campaign: dict[str, Any], day: str) -> float:
    start_at = parse_local_datetime(campaign.get("intakeStartDateTime"))
    end_at = parse_local_datetime(campaign.get("intakeEndDateTime"))
    if not start_at or not end_at:
        return 1.0 if campaign.get("intakeStart") <= day <= campaign.get("intakeEnd") else 0.0
    day_start = parse_local_datetime(f"{day}T00:00")
    day_end = parse_local_datetime(f"{day}T24:00")
    if not day_start or not day_end:
        return 0.0
    overlap_seconds = max(0.0, (min(end_at, day_end) - max(start_at, day_start)).total_seconds())
    return overlap_seconds / 86400


def campaign_target_share_map(campaigns: list[dict[str, Any]]) -> dict[str, float]:
    groups: dict[str, int] = defaultdict(int)
    for campaign in campaigns:
        key = campaign.get("openDate") or campaign.get("name")
        groups[key] += 1
    return {
        campaign["name"]: 1 / max(groups.get(campaign.get("openDate") or campaign.get("name"), 1), 1)
        for campaign in campaigns
        if campaign.get("name")
    }


def campaign_touches_month(campaign: dict[str, Any], month: str) -> bool:
    if any(day.startswith(month) for day in daterange(campaign.get("intakeStart"), campaign.get("intakeEnd"))):
        return True
    return any(add_days(campaign.get("openDate"), int(stage[1:]) - 1).startswith(month) for stage in D_STAGES)


def month_column_name(month: str, suffix: str) -> str:
    return f"{int(month[-2:])}月{suffix}"


def copy_row_style(ws, source_row: int, target_row: int):
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def build_pony_budget_workbook(config: dict[str, Any], month: str) -> tuple[Path, dict[str, Any]]:
    if not re.match(r"^\d{4}-\d{2}$", month):
        raise ValueError("月份格式不正确，请使用 YYYY-MM。")
    if not PONY_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"缺少 Pony 模板文件：{PONY_TEMPLATE_PATH}")

    campaigns = sorted(config.get("campaigns", []), key=lambda x: (x.get("openDate") or "", x.get("name") or ""))
    subchannels = config.get("subchannels", [])
    sub_by_id = {item["id"]: item for item in subchannels}
    sub_order = {item["id"]: index for index, item in enumerate(subchannels)}
    targets = target_lookup(config)
    share_by_campaign = campaign_target_share_map(campaigns)

    rows: list[dict[str, Any]] = []
    for campaign in campaigns:
        if not campaign_touches_month(campaign, month):
            continue
        open_date = campaign.get("openDate") or ""
        conversion_dates = {stage: add_days(open_date, int(stage[1:]) - 1) for stage in D_STAGES}
        r_values = r_lookup(config, campaign)
        share = share_by_campaign.get(campaign.get("name"), 1)
        for sub_id in campaign.get("subchannelIds", []):
            leads = sum(
                targets.get((day, sub_id), 0.0) * target_weight_for_day(campaign, day)
                for day in daterange(campaign.get("intakeStart"), campaign.get("intakeEnd"))
            ) * share
            if leads <= 0:
                continue
            daily_rates = [r_values.get((sub_id, stage), 0.0) / PONY_PRODUCT_PRICE for stage in D_STAGES]
            month_conversion_rate = sum(
                daily_rates[index]
                for index, stage in enumerate(D_STAGES)
                if conversion_dates.get(stage, "").startswith(month)
            )
            sub = sub_by_id.get(sub_id, {"name": sub_id})
            rows.append({
                "_subOrder": sub_order.get(sub_id, 999),
                "_openDate": open_date,
                "渠道": sub["name"],
                "营期": campaign.get("name", ""),
                "sku": "书画",
                "转化产品单价": PONY_PRODUCT_PRICE,
                "接量时间": normalize_date(campaign.get("intakeStartDateTime") or campaign.get("intakeStart")) or "",
                "接量截止时间": normalize_date(campaign.get("intakeEndDateTime") or campaign.get("intakeEnd")) or "",
                "开课时间": open_date,
                "转化时间": add_days(open_date, 3),
                "封板时间": add_days(open_date, 12),
                "lead数量": round(leads),
                "leads单价": "",
                "转化率-汇总": sum(daily_rates),
                **{f"转化率-DAY{index + 1}": daily_rates[index] for index in range(10)},
                "当月转化率": month_conversion_rate,
                month_column_name(month, "对应GMV"): round(leads * PONY_PRODUCT_PRICE * month_conversion_rate, 2),
                month_column_name(month, "leads数量"): round(leads) if month_conversion_rate > 0 else 0,
                month_column_name(month, "对应消耗"): "",
            })
    rows.sort(key=lambda row: (row["_subOrder"], row["_openDate"], row["营期"]))

    wb = load_workbook(PONY_TEMPLATE_PATH)
    ws = wb["Sheet1"]
    ws["X3"] = month_column_name(month, "对应GMV")
    ws["Y3"] = month_column_name(month, "leads数量")
    ws["Z3"] = month_column_name(month, "对应消耗")
    headers = [ws.cell(3, col).value for col in range(1, ws.max_column + 1)]
    if ws.max_row > 3:
        ws.delete_rows(4, ws.max_row - 3)
    for index, row in enumerate(rows, start=4):
        copy_row_style(ws, 3, index)
        for col, header in enumerate(headers, start=1):
            if header:
                ws.cell(index, col).value = row.get(header, "")
        for col in range(5, 10):
            ws.cell(index, col).number_format = "yyyy-mm-dd hh:mm"
        for col in range(12, 24):
            ws.cell(index, col).number_format = "0.00%"
        for col in range(24, 27):
            ws.cell(index, col).number_format = "#,##0"

    totals = {
        "lead数量": sum(row["lead数量"] for row in rows),
        month_column_name(month, "对应GMV"): sum(row[month_column_name(month, "对应GMV")] for row in rows),
        month_column_name(month, "leads数量"): sum(row[month_column_name(month, "leads数量")] for row in rows),
    }
    ws["A1"] = totals["lead数量"]
    ws["B1"] = totals[month_column_name(month, "leads数量")]
    ws["C1"] = totals[month_column_name(month, "对应GMV")]
    ws["D1"] = ""

    EXPORT_DIR.mkdir(exist_ok=True)
    output_path = EXPORT_DIR / f"Pony表营期规划-系统预算填充版-{month}.xlsx"
    wb.save(output_path)
    return output_path, {"rows": len(rows), "totals": totals}


def campaign_targets(config: dict[str, Any], as_of: str | None = None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    targets = target_lookup(config)
    subchannels_by_id = {item["id"]: item for item in config.get("subchannels", [])}
    generated: list[dict[str, Any]] = []
    daily_totals: dict[str, dict[str, Any]] = defaultdict(lambda: {
        "targetLeads": 0.0,
        "targetGmv": 0.0,
        "subchannels": defaultdict(lambda: {"targetLeads": 0.0, "targetGmv": 0.0}),
    })

    for campaign in config.get("campaigns", []):
        sub_ids = [str(x) for x in campaign.get("subchannelIds", [])]
        intake_days = daterange(campaign.get("intakeStart"), campaign.get("intakeEnd"))
        open_date = campaign.get("openDate")
        stage = stage_for_day(open_date, as_of or open_date)
        r_values = r_lookup(config, campaign)
        sub_targets = []
        total_leads = 0.0

        for sub_id in sub_ids:
            leads = sum(targets.get((day, sub_id), 0.0) for day in intake_days)
            total_leads += leads
            r_value = r_values.get((sub_id, stage), 0.0) if stage in D_STAGES else 0.0
            sub_targets.append({
                "subchannelId": sub_id,
                "subchannelName": subchannels_by_id.get(sub_id, {}).get("name", sub_id),
                "targetLeads": leads,
                "targetR": r_value,
                "targetGmv": leads * r_value,
            })

        weighted_r = pct(sum(item["targetLeads"] * item["targetR"] for item in sub_targets), total_leads)
        target_gmv = total_leads * weighted_r if stage in D_STAGES else 0.0
        generated.append({
            **campaign,
            "stage": stage,
            "targetLeads": total_leads,
            "targetR": weighted_r,
            "targetGmv": target_gmv,
            "subTargets": sub_targets,
        })

        for day in daterange(open_date, (parse_iso_day(open_date) + timedelta(days=13)).isoformat() if parse_iso_day(open_date) else open_date):
            day_stage = stage_for_day(open_date, day)
            day_r_values = r_lookup(config, campaign)
            for sub_id in sub_ids:
                leads = sum(targets.get((intake_day, sub_id), 0.0) for intake_day in intake_days)
                target_r = day_r_values.get((sub_id, day_stage), 0.0) if day_stage in D_STAGES else 0.0
                target_gmv = leads * target_r
                daily_totals[day]["targetLeads"] += 0.0
                daily_totals[day]["targetGmv"] += target_gmv
                daily_totals[day]["subchannels"][sub_id]["targetGmv"] += target_gmv

        for day in intake_days:
            for sub_id in sub_ids:
                leads = targets.get((day, sub_id), 0.0)
                daily_totals[day]["targetLeads"] += leads
                daily_totals[day]["subchannels"][sub_id]["targetLeads"] += leads

    return generated, daily_totals


def cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def to_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if text.endswith("%"):
            try:
                return float(text[:-1]) / 100
            except ValueError:
                return 0.0
        try:
            return float(text)
        except ValueError:
            return 0.0
    return 0.0


def normalize_date(value: Any) -> str | None:
    value = cell_value(value)
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        base = date(1899, 12, 30)
        return date.fromordinal(base.toordinal() + int(value)).isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            pass
    return text[:10]


def pct(numerator: float, denominator: float) -> float:
    return numerator / denominator if denominator else 0.0


DEFAULT_TEAM_MAPPINGS = [
    {"ownerName": "汪国炳", "teamName": "BD1"},
    {"ownerName": "孙晓迪", "teamName": "BD2"},
    {"ownerName": "卢宁", "teamName": "APP"},
    {"ownerName": "杜思博", "teamName": "BD1"},
    {"ownerName": "李博伟", "teamName": "BD2"},
    {"ownerName": "齐海洋1", "teamName": "私域"},
    {"ownerName": "李聿为", "teamName": "BD1"},
    {"ownerName": "郭庆函", "teamName": "BD2"},
    {"ownerName": "张艺珂", "teamName": "BD2"},
    {"ownerName": "王阿芳-投放", "teamName": "私域"},
    {"ownerName": "赵智勇2", "teamName": "BD2"},
    {"ownerName": "李俊超", "teamName": "BD2"},
    {"ownerName": "李云霞", "teamName": "直播"},
    {"ownerName": "李振亮", "teamName": "BD2"},
    {"ownerName": "刘检华", "teamName": "BD1"},
    {"ownerName": "任磬语", "teamName": "待填写"},
    {"ownerName": "马宇2", "teamName": "私域"},
]

CHANNEL_ANALYSIS_FIELD_ALIASES = {
    "spend": ["校正金额", "消耗总金额"],
    "leads": ["leads数"],
    "wechatAdds": ["加微数"],
    "activeWechatAdds": ["主动加微数"],
    "groups": ["总进群数", "入群数"],
    "d1Attend": ["d1到课数", "D1到课"],
    "d1Complete": ["d1完课数", "D1完课"],
    "d4Attend": ["d4到课数", "D4到课"],
    "d4Complete": ["d4完课数", "D4完课"],
    "d8Attend": ["d8到课数", "D8到课"],
    "d8Complete": ["d8完课数", "D8完课"],
    "currentGmv": ["当期成交金额", "当期成交gmv"],
    "previousGmv": ["往期成交gmv"],
    "midCourseGmv": ["课中成交"],
    "afterCourseGmv": ["课下成交"],
    "gmv": ["当期成交金额", "总成交额"],
    "income": ["当期成交金额", "总收入额"],
    "channelIncome": ["渠道收入"],
    "totalCost": ["总成本"],
}


def valid_studio_name(value: str) -> str:
    text = str(value or "").strip()
    if not text or text in {"待填写", "未校准工作室", "未校准团队"}:
        return ""
    return text


def default_studio_for_owner(owner: str) -> str:
    mapping = next((item for item in DEFAULT_TEAM_MAPPINGS if item["ownerName"] == owner), None)
    return mapping["teamName"] if mapping else ""


def normalize_channel_platform(value: str) -> str:
    text = str(value or "").strip()
    if text == "其他":
        return "其他平台"
    return text or "未标平台"


def channel_business_category(book_type: str, channel_type: str, studio_name: str) -> str:
    book = str(book_type or "").strip()
    channel = str(channel_type or "").strip()
    studio = valid_studio_name(studio_name)
    if book == "图书":
        return "图书"
    if book == "非图书" and channel == "KOL":
        return "0元"
    if book == "非图书" and channel == "短信/cps":
        return "短信"
    return studio or "其他"


def channel_platform_segment(business_category: str, platform: str) -> str:
    platform_name = normalize_channel_platform(platform)
    if business_category == "图书":
        return f"{platform_name}图书"
    if business_category == "0元":
        return f"{platform_name}0元"
    if business_category == "短信":
        return "短信渠道"
    return business_category


def campaign_display_label(open_date: str, campaign_name: str) -> str:
    prefix = open_date[5:] if open_date else "--"
    return f"{prefix} · {campaign_name}"


def enrich_channel_analysis_metrics(item: dict[str, Any]) -> dict[str, Any]:
    leads = to_number(item.get("leads"))
    spend = to_number(item.get("spend"))
    income = to_number(item.get("income"))
    gmv = to_number(item.get("gmv"))
    item["cpl"] = pct(spend, leads)
    item["roi"] = pct(income, spend)
    item["incomePerLead"] = pct(income, leads)
    item["gmvPerLead"] = pct(gmv, leads)
    item["wechatRate"] = pct(to_number(item.get("wechatAdds")), leads)
    item["groupRate"] = pct(to_number(item.get("groups")), leads)
    item["d1AttendRate"] = pct(to_number(item.get("d1Attend")), leads)
    item["d4AttendRate"] = pct(to_number(item.get("d4Attend")), leads)
    item["d8AttendRate"] = pct(to_number(item.get("d8Attend")), leads)
    return item


def aggregate_channel_analysis(rows: list[dict[str, Any]], key_fields: list[str]) -> list[dict[str, Any]]:
    groups: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in rows:
        key = tuple(row.get(field, "") for field in key_fields)
        if key not in groups:
            groups[key] = {field: row.get(field, "") for field in key_fields}
            groups[key].update({
                "rows": 0,
                "campaigns": set(),
                "owners": set(),
                "channelIds": set(),
                "rawStudios": set(),
            })
            for metric in CHANNEL_ANALYSIS_FIELD_ALIASES:
                groups[key][metric] = 0.0
        group = groups[key]
        group["rows"] += 1
        group["campaigns"].add(row.get("campaignName", ""))
        group["owners"].add(row.get("owner", ""))
        group["channelIds"].add(row.get("channelId", ""))
        group["rawStudios"].add(row.get("rawStudio", ""))
        for metric in CHANNEL_ANALYSIS_FIELD_ALIASES:
            group[metric] += to_number(row.get(metric))
    result: list[dict[str, Any]] = []
    for group in groups.values():
        group["campaignCount"] = len([x for x in group.pop("campaigns") if x])
        group["ownerCount"] = len([x for x in group.pop("owners") if x])
        group["channelIdCount"] = len([x for x in group.pop("channelIds") if x])
        group["rawStudioDistribution"] = "、".join(sorted(x for x in group.pop("rawStudios") if x))
        result.append(enrich_channel_analysis_metrics(group))
    return result


def read_channel_analysis_source_rows(path: str, filename: str) -> tuple[list[dict[str, Any]], int]:
    if not filename.lower().endswith(".csv"):
        raise ValueError("投放分析仅支持新版 CSV 明细，请上传包含 平台、渠道、图书、渠道归属、渠道号 的 CSV 文件。")
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = [{str(key or "").strip(): value for key, value in row.items()} for row in reader]
        return rows, len(reader.fieldnames or [])


def first_source_value(source: dict[str, Any], fields: list[str]) -> Any:
    for field in fields:
        if field in source and str(source.get(field) or "").strip() != "":
            return source.get(field)
    return ""


def parse_channel_analysis_import(path: str, filename: str) -> dict[str, Any]:
    source_rows, header_count = read_channel_analysis_source_rows(path, filename)
    if not source_rows:
        raise ValueError("没有识别到可分析的数据行。")
    headers = set(source_rows[0].keys())
    required = {"营期", "开课日期", "平台", "渠道", "图书", "渠道归属", "渠道号", "leads数", "校正金额"}
    missing = sorted(required - headers)
    if missing:
        raise ValueError(f"没有识别到新版投放明细字段：{', '.join(missing)}。")
    rows: list[dict[str, Any]] = []
    for source in source_rows:
        campaign_name = str(source.get("营期") or "").strip()
        channel_id = str(source.get("渠道号") or "").strip()
        owner = str(source.get("渠道归属") or "").strip() or "未归属"
        if not campaign_name or not channel_id:
            continue
        open_date = normalize_date(source.get("开课日期")) or ""
        platform = str(source.get("平台") or "").strip()
        channel = str(source.get("渠道") or "").strip() or "未分类"
        book_type = str(source.get("图书") or "").strip() or "未填写"
        studio = default_studio_for_owner(owner)
        business_category = channel_business_category(book_type, channel, studio)
        platform_segment = channel_platform_segment(business_category, platform)
        row = {
            "campaignName": campaign_name,
            "openDate": open_date,
            "dateCalibrated": bool(open_date),
            "campaignLabel": campaign_display_label(open_date, campaign_name),
            "leadDate": normalize_date(source.get("线索进量日期")) or "",
            "closeDate": normalize_date(source.get("封板日期")) or "",
            "dayLabel": str(source.get("昨天Day几") or "").strip(),
            "platform": platform or "未填写",
            "channel": channel,
            "channelId": channel_id,
            "businessCategory": business_category,
            "platformSegment": platform_segment,
            "bookType": book_type,
            "productType": book_type,
            "owner": owner,
            "studioName": studio if valid_studio_name(studio) else "未校准工作室",
            "teamName": studio if valid_studio_name(studio) else "未校准工作室",
            "rawStudio": str(source.get("供应商归属") or "").strip() or "未填写",
            "supplierTag": str(source.get("供应商标签") or "").strip(),
            "landingPageType": str(source.get("落地页类型") or "").strip(),
        }
        for metric, fields in CHANNEL_ANALYSIS_FIELD_ALIASES.items():
            row[metric] = to_number(first_source_value(source, fields))
        rows.append(enrich_channel_analysis_metrics(row))
    if not rows:
        raise ValueError("没有识别到可分析的渠道号明细。")

    summary = aggregate_channel_analysis(rows, [])[0]
    summary.update({
        "rowCount": len(rows),
        "campaignCount": len({row["campaignName"] for row in rows}),
        "ownerCount": len({row["owner"] for row in rows}),
        "channelIdCount": len({row["channelId"] for row in rows}),
        "columnCount": header_count,
        "uncalibratedOwnerCount": len({row["owner"] for row in rows if not valid_studio_name(default_studio_for_owner(row["owner"]))}),
    })
    campaigns = aggregate_channel_analysis(rows, ["campaignName", "openDate", "campaignLabel", "dateCalibrated"])
    owners = aggregate_channel_analysis(rows, ["owner"])
    studios = aggregate_channel_analysis(rows, ["studioName"])
    business_categories = aggregate_channel_analysis(rows, ["businessCategory"])
    platform_segments = aggregate_channel_analysis(rows, ["businessCategory", "platformSegment"])
    channel_ids = aggregate_channel_analysis(rows, ["studioName", "owner", "businessCategory", "platformSegment", "channelId", "channel", "bookType", "rawStudio"])
    campaigns.sort(key=lambda item: (str(item.get("openDate") or ""), str(item.get("campaignName") or "")))
    owners.sort(key=lambda item: (-to_number(item.get("income")), str(item.get("owner") or "")))
    studios.sort(key=lambda item: (-to_number(item.get("leads")), str(item.get("studioName") or "")))
    business_categories.sort(key=lambda item: (-to_number(item.get("leads")), str(item.get("businessCategory") or "")))
    platform_segments.sort(key=lambda item: (str(item.get("businessCategory") or ""), -to_number(item.get("leads"))))
    channel_ids.sort(key=lambda item: (str(item.get("studioName") or ""), str(item.get("owner") or ""), str(item.get("businessCategory") or ""), -to_number(item.get("income"))))
    return {
        "rows": rows,
        "summary": summary,
        "campaigns": campaigns,
        "owners": owners,
        "studios": studios,
        "businessCategories": business_categories,
        "platformSegments": platform_segments,
        "channelIds": channel_ids,
    }


def status_for_rate(rate: float, good_low: float = 0.9, good_high: float = 1.1) -> str:
    if rate <= 0:
        return "empty"
    if rate < 0.8 or rate > 1.2:
        return "danger"
    if good_low <= rate <= good_high:
        return "good"
    return "warn"


def first_sheet(workbook, candidates: list[str]):
    for name in candidates:
        if name in workbook.sheetnames:
            return workbook[name]
    return None


def headers(ws, row: int = 1) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row, col).value
        if value is not None and str(value).strip():
            result[str(value).strip()] = col
    return result


def parse_actual_daily(workbook) -> tuple[list[dict[str, Any]], dict[str, dict[str, float]]]:
    ws = first_sheet(workbook, ["【数据】实际-每日GMV（下载覆盖）"])
    if not ws:
        return [], {}

    daily: list[dict[str, Any]] = []
    by_date: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in range(2, ws.max_row + 1):
        sku = str(ws.cell(row, 1).value or "").strip()
        raw_date = ws.cell(row, 2).value
        category = str(ws.cell(row, 3).value or "").strip()
        if sku != "日-汇总":
            continue
        day = normalize_date(raw_date)
        if not day:
            continue
        item = {
            "date": day,
            "category": category,
            "actualLeads": to_number(ws.cell(row, 4).value),
            "fullPriceStudents": to_number(ws.cell(row, 5).value),
            "wechatAdds": to_number(ws.cell(row, 6).value),
            "wechatRate": to_number(ws.cell(row, 7).value),
            "actualGmv": to_number(ws.cell(row, 8).value),
            "spend": to_number(ws.cell(row, 16).value),
            "crmSpend": to_number(ws.cell(row, 17).value),
        }
        daily.append(item)
        bucket = by_date[day]
        for key in ("actualLeads", "fullPriceStudents", "wechatAdds", "actualGmv", "spend", "crmSpend"):
            bucket[key] += float(item[key])
    daily.sort(key=lambda x: x["date"])
    return daily, by_date


def parse_actual_daily_by_subchannel(workbook) -> dict[str, dict[str, dict[str, float]]]:
    ws = first_sheet(workbook, ["【数据】实际-每日GMV（下载覆盖）"])
    if not ws:
        return {}
    by_day: dict[str, dict[str, dict[str, float]]] = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    ignored_markers = {"团队-汇总", "萱哥要的分类", "分类-汇总", "sku-汇总"}
    for row in range(2, ws.max_row + 1):
        raw_date = ws.cell(row, 2).value
        day = normalize_date(raw_date)
        sku = str(ws.cell(row, 1).value or "").strip()
        category = str(ws.cell(row, 3).value or "").strip()
        if not day or sku == "日-汇总" or str(raw_date).strip() in ignored_markers:
            continue
        if not re.match(r"\d{4}-\d{2}-\d{2}", day):
            continue
        bucket = by_day[day][category or "-"]
        bucket["actualLeads"] += to_number(ws.cell(row, 4).value)
        bucket["fullPriceStudents"] += to_number(ws.cell(row, 5).value)
        bucket["actualGmv"] += to_number(ws.cell(row, 8).value)
        bucket["spend"] += to_number(ws.cell(row, 16).value)
    return by_day


def parse_targets(workbook) -> tuple[dict[str, float], dict[str, float], dict[str, list[dict[str, Any]]]]:
    schedule = first_sheet(workbook, ["排期表（日）"])
    revenue = first_sheet(workbook, ["营收表（GMV）"])
    target_leads: dict[str, float] = {}
    target_gmv: dict[str, float] = {}
    stages: dict[str, list[dict[str, Any]]] = defaultdict(list)

    if schedule:
        for row in range(4, schedule.max_row + 1):
            day = normalize_date(schedule.cell(row, 2).value)
            if not day:
                continue
            target_leads[day] = to_number(schedule.cell(row, 3).value)
            for col in range(4, schedule.max_column + 1):
                stage = schedule.cell(row, col).value
                if stage is None or str(stage).strip() == "":
                    continue
                code = schedule.cell(1, col).value
                name = schedule.cell(2, col).value or f"{code}期"
                stages[day].append({
                    "periodCode": str(code or "").strip(),
                    "periodName": str(name).strip(),
                    "stage": str(stage).strip(),
                })

    if revenue:
        for row in range(3, revenue.max_row + 1):
            day = normalize_date(revenue.cell(row, 2).value)
            if not day:
                continue
            target_gmv[day] = to_number(revenue.cell(row, 3).value)

    return target_leads, target_gmv, stages


def record_from_row(ws, row: int, header_map: dict[str, int]) -> dict[str, Any]:
    item: dict[str, Any] = {}
    for key, col in header_map.items():
        item[key] = cell_value(ws.cell(row, col).value)
    return item


def period_code(name: str) -> str:
    match = re.search(r"(\d{4})$", name or "")
    return match.group(1) if match else ""


def parse_campaigns(
    workbook,
    stages_by_date: dict[str, list[dict[str, Any]]],
    stage_as_of: str | None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    ws = first_sheet(workbook, ["【数据】分营期-过程&结果数据"])
    if not ws:
        return [], [], []

    header_map = headers(ws, 1)
    r_fields = [
        name for name in header_map
        if re.search(r"(?:^D\d+|R值)", name, re.IGNORECASE)
    ]

    raw_records = [record_from_row(ws, row, header_map) for row in range(2, ws.max_row + 1)]
    detail_rows = [
        row for row in raw_records
        if str(row.get("营期") or "").strip()
        and str(row.get("营期")).strip() not in ("分类总计", "分类合计")
    ]
    category_rows = [
        row for row in raw_records
        if str(row.get("营期") or "").strip() in ("分类合计", "分类总计")
    ]

    latest_stage_by_code: dict[str, str] = {}
    stage_days = [day for day in sorted(stages_by_date) if not stage_as_of or day <= stage_as_of]
    if not stage_days:
        stage_days = sorted(stages_by_date)
    for day in stage_days:
        for stage in stages_by_date[day]:
            code = str(stage.get("periodCode") or "")
            if code:
                latest_stage_by_code[code] = str(stage.get("stage") or "")

    campaigns_by_name: dict[str, dict[str, Any]] = {}
    for row in detail_rows:
        name = str(row.get("营期") or "").strip()
        if not name:
            continue
        bucket = campaigns_by_name.setdefault(name, {
            "name": name,
            "openDate": normalize_date(row.get("开课日期")),
            "closeDate": normalize_date(row.get("封板日期")),
            "owner": row.get("渠道归属人") or "",
            "categories": set(),
            "leads": 0.0,
            "gmv": 0.0,
            "spend": 0.0,
            "fullPriceStudents": 0.0,
            "wechatAdds": 0.0,
            "rBreakdown": defaultdict(float),
        })
        bucket["categories"].add(str(row.get("分类") or "").strip())
        bucket["leads"] += to_number(row.get("leads数"))
        bucket["gmv"] += to_number(row.get("gmv"))
        bucket["spend"] += to_number(row.get("消耗"))
        bucket["fullPriceStudents"] += to_number(row.get("正价课学员数") or row.get("正价学员数"))
        bucket["wechatAdds"] += to_number(row.get("加微数"))
        for field in r_fields:
            value = to_number(row.get(field))
            if value:
                bucket["rBreakdown"][field] += value

    campaigns: list[dict[str, Any]] = []
    for item in campaigns_by_name.values():
        code = period_code(item["name"])
        r_breakdown = dict(item["rBreakdown"])
        r_total = sum(r_breakdown.values())
        item["periodCode"] = code
        item["stage"] = latest_stage_by_code.get(code, "未匹配")
        item["categories"] = "、".join(sorted(x for x in item["categories"] if x))
        item["conversionRate"] = pct(item["fullPriceStudents"], item["leads"])
        item["roi"] = pct(item["gmv"], item["spend"])
        item["rValue"] = pct(item["gmv"], item["leads"])
        item["rBreakdown"] = r_breakdown
        campaigns.append(item)

    channels: list[dict[str, Any]] = []
    for row in category_rows:
        leads = to_number(row.get("leads数"))
        gmv = to_number(row.get("gmv"))
        spend = to_number(row.get("消耗"))
        full_students = to_number(row.get("正价课学员数") or row.get("正价学员数"))
        r_breakdown = {field: to_number(row.get(field)) for field in r_fields if to_number(row.get(field))}
        channels.append({
            "scope": row.get("营期"),
            "category": row.get("分类") or "-",
            "leads": leads,
            "gmv": gmv,
            "spend": spend,
            "fullPriceStudents": full_students,
            "conversionRate": pct(full_students, leads),
            "roi": pct(gmv, spend),
            "rValue": pct(gmv, leads),
            "rBreakdown": r_breakdown,
        })

    campaigns.sort(key=lambda x: (x.get("openDate") or "", x["name"]), reverse=True)
    channels.sort(key=lambda x: x["roi"])
    return campaigns, channels, r_fields


def build_warnings(
    daily: list[dict[str, Any]],
    campaigns: list[dict[str, Any]],
    channels: list[dict[str, Any]],
    latest_actual_date: str | None = None,
) -> list[dict[str, Any]]:
    warnings: list[dict[str, Any]] = []
    for day in daily:
        if latest_actual_date and day["date"] > latest_actual_date:
            continue
        if day["targetLeads"] and day["leadsRateStatus"] == "danger":
            warnings.append({
                "type": "Leads",
                "level": "danger",
                "title": f"{day['date']} Leads 达成异常",
                "detail": f"目标 {day['targetLeads']:.0f}，实际 {day['actualLeads']:.0f}，达成率 {day['leadsRate']:.0%}",
            })
        if day["targetGmv"] and day["gmvRate"] < 0.8:
            warnings.append({
                "type": "GMV",
                "level": "danger",
                "title": f"{day['date']} GMV 达成不足",
                "detail": f"目标 {day['targetGmv']:.0f}，实际 {day['actualGmv']:.0f}，达成率 {day['gmvRate']:.0%}",
            })
    for item in campaigns:
        if item["leads"] and item["gmv"] == 0:
            warnings.append({
                "type": "营期",
                "level": "warn",
                "title": f"{item['name']} 暂无 GMV",
                "detail": f"当前阶段 {item['stage']}，Leads {item['leads']:.0f}，需检查转化承接。",
            })
    for item in channels[:8]:
        if item["spend"] and item["roi"] < 1:
            warnings.append({
                "type": "渠道",
                "level": "warn",
                "title": f"{item['category']} ROI 偏低",
                "detail": f"ROI {item['roi']:.2f}，GMV {item['gmv']:.0f}，消耗 {item['spend']:.0f}",
            })
    return warnings[:30]


def merge_campaign_targets(actual_campaigns: list[dict[str, Any]], target_campaigns: list[dict[str, Any]]) -> list[dict[str, Any]]:
    actual_by_name = {item["name"]: item for item in actual_campaigns}
    merged: list[dict[str, Any]] = []
    for target in target_campaigns:
        actual = actual_by_name.pop(target["name"], {})
        leads = to_number(actual.get("leads"))
        gmv = to_number(actual.get("gmv"))
        spend = to_number(actual.get("spend"))
        full_students = to_number(actual.get("fullPriceStudents"))
        target_leads = to_number(target.get("targetLeads"))
        target_gmv = to_number(target.get("targetGmv"))
        merged.append({
            **target,
            "actualLeads": leads,
            "actualGmv": gmv,
            "spend": spend,
            "fullPriceStudents": full_students,
            "categories": actual.get("categories") or "、".join(x.get("subchannelName", "") for x in target.get("subTargets", [])),
            "leads": leads or target_leads,
            "gmv": gmv,
            "conversionRate": pct(full_students, leads),
            "roi": pct(gmv, spend),
            "rValue": pct(gmv, leads),
            "targetLeadsRate": pct(leads, target_leads),
            "targetGmvRate": pct(gmv, target_gmv),
            "rBreakdown": actual.get("rBreakdown", {}),
        })
    for actual in actual_by_name.values():
        actual.setdefault("actualLeads", actual.get("leads", 0))
        actual.setdefault("actualGmv", actual.get("gmv", 0))
        actual.setdefault("targetLeads", 0)
        actual.setdefault("targetGmv", 0)
        actual.setdefault("targetR", 0)
        actual.setdefault("targetLeadsRate", 0)
        actual.setdefault("targetGmvRate", 0)
        merged.append(actual)
    merged.sort(key=lambda x: (x.get("openDate") or "", x.get("name", "")), reverse=True)
    return merged


def parse_workbook(path: str, config: dict[str, Any] | None = None) -> dict[str, Any]:
    config = config or read_config()
    workbook = load_workbook(path, data_only=True)
    actual_daily, actual_by_date = parse_actual_daily(workbook)
    actual_by_subchannel = parse_actual_daily_by_subchannel(workbook)
    target_leads, target_gmv, stages_by_date = parse_targets(workbook)
    all_dates = sorted(set(target_leads) | set(target_gmv) | set(actual_by_date))
    latest_actual_date = max((day for day, data in actual_by_date.items() if data.get("actualLeads") or data.get("actualGmv")), default=(all_dates[-1] if all_dates else None))
    actual_campaigns, channels, r_fields = parse_campaigns(workbook, stages_by_date, latest_actual_date)
    generated_campaigns, generated_daily = campaign_targets(config, latest_actual_date)
    campaigns = merge_campaign_targets(actual_campaigns, generated_campaigns) if generated_campaigns else actual_campaigns

    if generated_daily:
        target_leads = {day: to_number(data.get("targetLeads")) for day, data in generated_daily.items()}
        target_gmv = {day: to_number(data.get("targetGmv")) for day, data in generated_daily.items()}
        all_dates = sorted(set(target_leads) | set(target_gmv) | set(actual_by_date))

    daily: list[dict[str, Any]] = []
    for day in all_dates:
        actual = actual_by_date.get(day, {})
        subchannels = []
        generated_subchannels = generated_daily.get(day, {}).get("subchannels", {}) if generated_daily else {}
        for sub_id, values in generated_subchannels.items():
            sub_meta = next((x for x in config.get("subchannels", []) if x.get("id") == sub_id), {})
            sub_name = sub_meta.get("name", sub_id)
            actual_sub = actual_by_subchannel.get(day, {}).get(sub_name, {})
            sub_target_leads = to_number(values.get("targetLeads"))
            sub_actual_leads = to_number(actual_sub.get("actualLeads"))
            sub_target_gmv = to_number(values.get("targetGmv"))
            sub_actual_gmv = to_number(actual_sub.get("actualGmv"))
            subchannels.append({
                "subchannelId": sub_id,
                "subchannelName": sub_name,
                "targetLeads": sub_target_leads,
                "actualLeads": sub_actual_leads,
                "leadsRate": pct(sub_actual_leads, sub_target_leads),
                "targetGmv": sub_target_gmv,
                "actualGmv": sub_actual_gmv,
                "gmvRate": pct(sub_actual_gmv, sub_target_gmv),
            })
        row = {
            "date": day,
            "targetLeads": target_leads.get(day, 0.0),
            "actualLeads": float(actual.get("actualLeads", 0.0)),
            "targetGmv": target_gmv.get(day, 0.0),
            "actualGmv": float(actual.get("actualGmv", 0.0)),
            "fullPriceStudents": float(actual.get("fullPriceStudents", 0.0)),
            "stages": stages_by_date.get(day, []),
            "subchannels": subchannels,
        }
        row["leadsRate"] = pct(row["actualLeads"], row["targetLeads"])
        row["gmvRate"] = pct(row["actualGmv"], row["targetGmv"])
        row["leadsRateStatus"] = status_for_rate(row["leadsRate"])
        row["gmvRateStatus"] = "danger" if row["targetGmv"] and row["gmvRate"] < 0.8 else ("good" if row["gmvRate"] >= 1 else "warn")
        daily.append(row)

    today = next((d for d in daily if d["date"] == latest_actual_date), daily[-1] if daily else {})
    month_prefix = latest_actual_date[:7] if latest_actual_date else ""
    month_rows = [d for d in daily if d["date"].startswith(month_prefix)]
    month_target_gmv = sum(d["targetGmv"] for d in month_rows)
    month_actual_gmv = sum(d["actualGmv"] for d in month_rows)

    overview = {
        "latestDate": latest_actual_date,
        "todayTargetLeads": today.get("targetLeads", 0),
        "todayActualLeads": today.get("actualLeads", 0),
        "todayLeadsRate": today.get("leadsRate", 0),
        "todayTargetGmv": today.get("targetGmv", 0),
        "todayActualGmv": today.get("actualGmv", 0),
        "todayGmvRate": today.get("gmvRate", 0),
        "monthTargetGmv": month_target_gmv,
        "monthActualGmv": month_actual_gmv,
        "monthGmvRate": pct(month_actual_gmv, month_target_gmv),
        "campaignCount": len(campaigns),
        "channelCount": len(channels),
    }

    warnings = build_warnings(daily, campaigns, channels, latest_actual_date)
    overview["warningCount"] = len(warnings)

    return {
        "overview": overview,
        "daily": daily,
        "campaigns": campaigns,
        "channels": channels,
        "warnings": warnings,
        "rFields": r_fields,
        "config": config,
        "generatedCampaigns": generated_campaigns,
        "sheets": workbook.sheetnames,
    }


class DashboardHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        parsed_url = urlparse(self.path)
        request_path = parsed_url.path
        if request_path == "/api/health":
            self.send_json({
                "ok": True,
                "pid": os.getpid(),
                "host": os.environ.get("HOST", "127.0.0.1"),
                "port": int(os.environ.get("PORT", "8765")),
                "root": str(ROOT),
                "startedAt": SERVER_START_TIME,
            })
            return
        if request_path == "/api/config":
            self.send_json(read_config())
            return
        if request_path == "/api/revenue-overview":
            try:
                query = parse_qs(parsed_url.query)
                month = (query.get("month") or [""])[0] or datetime.now().strftime("%Y-%m")
                self.send_json(revenue_overview(month))
            except Exception as exc:
                self.send_json({"error": f"营收汇总失败：{type(exc).__name__}: {exc}"}, status=500)
            return
        if request_path == "/api/feishu-report-targets":
            try:
                self.send_json({"targets": feishu_report_targets()})
            except Exception as exc:
                self.send_json({"error": f"读取飞书接收群失败：{type(exc).__name__}: {exc}"}, status=500)
            return
        if request_path == "/api/export-pony-budget":
            try:
                query = parse_qs(parsed_url.query)
                month = (query.get("month") or [""])[0] or datetime.now().strftime("%Y-%m")
                output_path, meta = build_pony_budget_workbook(read_config(), month)
                body = output_path.read_bytes()
                filename = output_path.name
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(filename)}")
                self.send_header("Content-Length", str(len(body)))
                self.send_header("X-Pony-Rows", str(meta["rows"]))
                self.end_headers()
                self.wfile.write(body)
            except Exception as exc:
                self.send_json({"error": f"生成失败：{type(exc).__name__}: {exc}"}, status=500)
            return
        if request_path == "/api/generated-targets":
            config = read_config()
            campaigns, daily = campaign_targets(config, datetime.now().date().isoformat())
            serial_daily = {
                day: {
                    "targetLeads": data["targetLeads"],
                    "targetGmv": data["targetGmv"],
                    "subchannels": dict(data["subchannels"]),
                }
                for day, data in daily.items()
            }
            self.send_json({"campaigns": campaigns, "daily": serial_daily})
            return
        if request_path == "/api/operation-log":
            self.send_json({"logs": read_operation_log()})
            return
        if request_path == "/":
            request_path = "/index.html"
        file_path = (PUBLIC / request_path.lstrip("/")).resolve()
        if not str(file_path).startswith(str(PUBLIC.resolve())) or not file_path.exists():
            self.send_error(404)
            return
        content_type = "text/html"
        if file_path.suffix == ".css":
            content_type = "text/css"
        elif file_path.suffix == ".js":
            content_type = "application/javascript"
        self.send_response(200)
        self.send_header("Content-Type", f"{content_type}; charset=utf-8")
        self.end_headers()
        self.wfile.write(file_path.read_bytes())

    def do_POST(self):
        if self.path == "/api/feishu-card-callback":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
                self.send_json(handle_feishu_card_callback(payload))
            except Exception as exc:
                self.send_json({"error": f"飞书卡片回调失败：{type(exc).__name__}: {exc}"}, status=500)
            return

        if self.path == "/api/operation-log":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                append_operation_log({
                    **payload,
                    "client": self.client_address[0],
                    "userAgent": self.headers.get("User-Agent", ""),
                })
                self.send_json({"ok": True})
            except Exception as exc:
                self.send_json({"error": f"记录失败：{type(exc).__name__}: {exc}"}, status=500)
            return

        if self.path == "/api/revenue-sync":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
                month = payload.get("month") or datetime.now().strftime("%Y-%m")
                end_date = payload.get("endDate") or ""
                self.send_json(sync_revenue_actuals_from_crm(str(month), str(end_date)))
            except Exception as exc:
                self.send_json({"error": f"CRM 更新失败：{type(exc).__name__}: {exc}"}, status=500)
            return

        if self.path == "/api/feishu-report":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
                month = payload.get("month") or datetime.now().strftime("%Y-%m")
                scope = str(payload.get("scope") or "overall")
                target_id = str(payload.get("targetId") or "")
                self.send_json(send_feishu_overview_report(str(month), scope, target_id))
            except Exception as exc:
                self.send_json({"error": f"飞书日报发送失败：{type(exc).__name__}: {exc}"}, status=500)
            return

        if self.path == "/api/config":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                write_config(payload)
                self.send_json(payload)
            except Exception as exc:
                self.send_json({"error": f"保存失败：{type(exc).__name__}: {exc}"}, status=500)
            return

        if self.path not in {"/api/upload", "/api/import-actual", "/api/import-channel-analysis"}:
            self.send_error(404)
            return
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type"),
            },
        )
        upload = form["file"] if "file" in form else None
        if upload is None or not upload.file:
            self.send_json({"error": "没有收到文件"}, status=400)
            return
        filename = getattr(upload, "filename", "") or "upload.xlsx"
        try:
            suffix = Path(filename).suffix or ".xlsx"
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=True) as tmp:
                tmp.write(upload.file.read())
                tmp.flush()
                if self.path == "/api/import-actual":
                    actuals = parse_actual_import(tmp.name, filename)
                    parsed = {"actualCampaigns": actuals}
                elif self.path == "/api/import-channel-analysis":
                    parsed = parse_channel_analysis_import(tmp.name, filename)
                else:
                    parsed = parse_workbook(tmp.name, read_config())
            self.send_json(parsed)
        except Exception as exc:  # Keep API visible for first-version local tool.
            self.send_json({"error": f"解析失败：{type(exc).__name__}: {exc}"}, status=500)

    def send_json(self, payload: dict[str, Any], status: int = 200):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


if __name__ == "__main__":
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "8765"))
    server = ThreadingHTTPServer((host, port), DashboardHandler)
    print(f"营期转化监控看板已启动：http://{host}:{port}", flush=True)
    if host == "0.0.0.0":
        print(f"局域网访问地址：http://{local_lan_ip()}:{port}", flush=True)
        print("同一个 Wi-Fi / 局域网下的电脑或手机可以打开上面的地址。", flush=True)
    server.serve_forever()
